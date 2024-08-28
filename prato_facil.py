import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

# Carregar as tabelas
tabela1 = pd.read_csv('Beneficiarios Prato Fácil.csv', delimiter=';')
tabela2 = pd.read_csv('prato.csv', delimiter=';')

# Filtrar apenas os valores que são numéricos na coluna NIS da Tabela 1
tabela1 = tabela1[tabela1['NIS'].str.isnumeric()]

# Converter a coluna NIS da Tabela 1 para float64
tabela1['NIS'] = tabela1['NIS'].astype(float)
tabela2['NIS'] = tabela2['NIS'].astype(float)

# Obter a contagem total de registros na Tabela 1
total_tabela1 = len(tabela1)

# Verificar quantos da Tabela 1 estão fora da Tabela 2
nis_tabela1 = tabela1['NIS']
nis_tabela2 = tabela2['NIS']

fora_da_tabela2 = nis_tabela1[~nis_tabela1.isin(nis_tabela2)]
num_fora_tabela2 = len(fora_da_tabela2)

# Fazer o merge das tabelas com base na coluna NIS
merged = pd.merge(tabela1, tabela2, on='NIS', how='inner')

# Filtrar por faixa salarial
meio_salario = merged[merged['d.vlr_renda_media_fam'] <= 706].shape[0]
um_salario = merged[(merged['d.vlr_renda_media_fam'] > 706) & (merged['d.vlr_renda_media_fam'] <= 1412)].shape[0]
mais_de_um_salario = merged[merged['d.vlr_renda_media_fam'] > 1412].shape[0]

# Mostrar resultados na tela
print(f"Total de registros na Tabela 1: {total_tabela1}")
print(f"Número de registros fora da Tabela 2: {num_fora_tabela2}")
print(f"Distribuição por faixa salarial:")
print(f"  - Até meio salário mínimo (706 reais): {meio_salario}")
print(f"  - Até 1 salário mínimo (1412 reais): {um_salario}")
print(f"  - Mais de 1 salário mínimo: {mais_de_um_salario}")

# Salvar os resultados em um arquivo Excel
wb = Workbook()
ws = wb.active
ws.title = "Resultados"

# Adicionar a contagem total da Tabela 1 ao Excel
ws.cell(row=1, column=1, value="TOTAL DE REGISTROS BENEFICIÁRIOS DO PRATO FÁCIL")
ws.cell(row=2, column=1, value=total_tabela1)

# Adicionar os outros resultados ao Excel
ws.cell(row=4, column=1, value="NÚMERO DE REGISTROS FORA DO CAD ÚNICO")
ws.cell(row=5, column=1, value=num_fora_tabela2)

ws.cell(row=7, column=1, value="DISTRIBUIÇÃO POR FAIXA SALARIAL")
ws.cell(row=8, column=1, value="Até meio salário mínimo (706 reais)")
ws.cell(row=8, column=2, value=meio_salario)
ws.cell(row=9, column=1, value="Até 1 salário mínimo (1412 reais)")
ws.cell(row=9, column=2, value=um_salario)
ws.cell(row=10, column=1, value="Mais de 1 salário mínimo (1412 reais)")
ws.cell(row=10, column=2, value=mais_de_um_salario)

# Formatar o cabeçalho em negrito
for cell in ws["1:1"] + ws["4:4"] + ws["7:7"]:
    cell.font = Font(bold=True)

# Salvar o arquivo Excel
wb.save("resultados_salariais.xlsx")

print("Resultados salvos em 'resultados_salariais.xlsx'")
