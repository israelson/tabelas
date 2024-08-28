import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

# Carregar os CSVs
tabela1 = pd.read_csv('Beneficiarios Prato Fácil.csv', delimiter=';')
tabela2 = pd.read_csv('vencer.csv', delimiter=';')

# Garantir que a coluna NIS está como string em ambas as tabelas
tabela1['NIS'] = tabela1['NIS'].astype(str)
tabela2['NIS'] = tabela2['NIS'].astype(str)

# Fazer o merge das tabelas com base na coluna NIS, mantendo apenas os registros que estão na tabela 1
merged = pd.merge(tabela1, tabela2, on='NIS', how='inner')

# Mapear valores de sexo e raça para suas descrições
sexo_map = {1: 'Masculino', 2: 'Feminino'}
raca_map = {1: 'Branca', 2: 'Preta', 3: 'Amarela', 4: 'Parda', 5: 'Indígena'}

# Aplicar mapeamento
merged['p.cod_sexo_pessoa'] = merged['p.cod_sexo_pessoa'].map(sexo_map)
merged['p.cod_raca_cor_pessoa'] = merged['p.cod_raca_cor_pessoa'].map(raca_map)

# Contagem de sexo
sexo_contagem = merged['p.cod_sexo_pessoa'].value_counts()

# Calcular a idade considerando o formato dia/mês/ano
merged['idade'] = pd.to_datetime(merged['p.dta_nasc_pessoa'], dayfirst=True).apply(lambda x: pd.Timestamp('now').year - x.year)

# Faixas etárias
faixa_etaria = {
    'Menor de 20': merged[merged['idade'] < 20].shape[0],
    '21 a 30': merged[(merged['idade'] >= 21) & (merged['idade'] <= 30)].shape[0],
    '31 a 40': merged[(merged['idade'] >= 31) & (merged['idade'] <= 40)].shape[0],
    '41 a 50': merged[(merged['idade'] >= 41) & (merged['idade'] <= 50)].shape[0],
    'Maior de 50': merged[merged['idade'] > 50].shape[0]
}

# Contagem por raça
raca_contagem = merged['p.cod_raca_cor_pessoa'].value_counts()

# Faixas salariais
salario_faixas = {
    'Até 706': merged[merged['d.vlr_renda_media_fam'] <= 700].shape[0],
    'Até 1412': merged[(merged['d.vlr_renda_media_fam'] > 700) & (merged['d.vlr_renda_media_fam'] <= 1400)].shape[0],
    'Mais de 1412': merged[merged['d.vlr_renda_media_fam'] > 1400].shape[0]
}

# Contagem por bairro
bairro_contagem = merged['d.nom_localidade_fam'].value_counts()

# Criar um arquivo Excel com os dados
wb = Workbook()
ws = wb.active
ws.title = "Resultados"

# Adicionar dados ao Excel organizados por seções
sections = [
    ("SEXO", sexo_contagem),
    ("RAÇA", raca_contagem),
    ("FAIXA ETÁRIA", pd.Series(faixa_etaria)),
    ("FAIXA SALARIAL", pd.Series(salario_faixas)),
    ("BAIRRO", bairro_contagem)
]

# Adicionar cada seção ao Excel
current_row = 1
for title, data in sections:
    ws.cell(row=current_row, column=1, value=title)
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    current_row += 1
    for index, value in data.items():
        ws.cell(row=current_row, column=1, value=index)
        ws.cell(row=current_row, column=2, value=value)
        current_row += 1
    current_row += 1  # Espaço entre seções

# Salvar o arquivo Excel
wb.save("resultados_formatados.xlsx")

print("Arquivo 'resultados_formatados.xlsx' foi salvo com sucesso!")
